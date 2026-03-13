"""Export wine_catalog.json to wine_catalog.xlsx (products + winery sheets).

Use this to create or refresh the Excel source from existing JSON.
After running, use import_wine_catalog.py to sync xlsx -> json (or use watch).

Usage:
  python scripts/export_wine_catalog.py
  python scripts/export_wine_catalog.py --input ../../00_docs/docs_md/spec/fastpath/wine_catalog.json --output ../../00_docs/docs_md/spec/fastpath/wine_catalog.xlsx
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook

# Columns for products sheet (first sheet) - order and names must match import
PRODUCT_HEADERS = [
    "id", "name", "name_cn", "sku", "color", "region", "grape_variety", "price",
    "acid", "tannin", "body", "sweetness", "flavor_profile", "food_pairing",
    "occasion", "tasting_notes", "alcohol", "vintage", "winery",
    "video_url", "video_title",
]
WINERY_SHEET_NAME = "winery"
WINERY_HEADERS = [
    "id", "name_zh", "name_en", "intro_zh", "intro_en",
    "selling_points_zh", "selling_points_en",
]


def load_json(json_path: Path) -> dict[str, Any]:
    data = json.loads(json_path.read_text(encoding="utf-8"))
    return {
        "products": data.get("products") or [],
        "wineries": data.get("wineries") or [],
    }


def write_xlsx(data: dict[str, Any], xlsx_path: Path) -> None:
    wb = Workbook()
    # Sheet 1: products
    ws1 = wb.active
    ws1.title = "products"
    for col, h in enumerate(PRODUCT_HEADERS, start=1):
        ws1.cell(row=1, column=col, value=h)
    for row_idx, prod in enumerate(data["products"], start=2):
        for col_idx, key in enumerate(PRODUCT_HEADERS, start=1):
            val = prod.get(key)
            if val is None and key in ("tannin", "video_url", "video_title"):
                val = ""
            ws1.cell(row=row_idx, column=col_idx, value=val)

    # Sheet 2: winery
    ws2 = wb.create_sheet(WINERY_SHEET_NAME)
    for col, h in enumerate(WINERY_HEADERS, start=1):
        ws2.cell(row=1, column=col, value=h)
    for row_idx, w in enumerate(data["wineries"], start=2):
        for col_idx, key in enumerate(WINERY_HEADERS, start=1):
            ws2.cell(row=row_idx, column=col_idx, value=w.get(key) or "")

    xlsx_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(xlsx_path)


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser()
    p.add_argument(
        "--input",
        type=Path,
        default=Path("../../00_docs/docs_md/spec/fastpath/wine_catalog.json"),
        help="Path to wine_catalog.json",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=Path("../../00_docs/docs_md/spec/fastpath/wine_catalog.xlsx"),
        help="Path to output wine_catalog.xlsx",
    )
    return p.parse_args()


def main() -> None:
    args = parse_args()
    base = Path(__file__).resolve().parent
    input_path = (base / args.input).resolve()
    output_path = (base / args.output).resolve()

    data = load_json(input_path)
    write_xlsx(data, output_path)
    print(f"Exported {len(data['products'])} product(s), {len(data['wineries'])} winery(ies): {input_path} -> {output_path}")


if __name__ == "__main__":
    main()
