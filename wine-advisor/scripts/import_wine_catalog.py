"""Import unified wine catalog from Excel into JSON.

Usage:
  python scripts/import_wine_catalog.py
  python scripts/import_wine_catalog.py --input ../../00_docs/docs_md/spec/fastpath/wine_catalog.xlsx --output ../../00_docs/docs_md/spec/fastpath/wine_catalog.json
"""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


REQUIRED_COLUMNS = [
    "id",
    "name",
    "name_cn",
    "sku",
    "color",
    "region",
    "grape_variety",
    "price",
    "acid",
    "tannin",
    "body",
    "sweetness",
    "flavor_profile",
    "food_pairing",
    "occasion",
    "tasting_notes",
    "alcohol",
    "vintage",
    "winery",
]

REQUIRED_NON_EMPTY_COLUMNS = [
    "id",
    "name",
    "name_cn",
    "sku",
    "color",
    "region",
    "grape_variety",
    "price",
    "acid",
    "body",
    "sweetness",
    "flavor_profile",
    "food_pairing",
    "occasion",
    "tasting_notes",
    "alcohol",
    "vintage",
    "winery",
]

OPTIONAL_COLUMNS = [
    "video_url",
    "video_title",
]

ALLOWED_COLORS = {"red", "white", "rose", "sparkling"}
ALLOWED_ACID = {"high", "moderate", "low"}
ALLOWED_TANNIN = {"strong", "medium", "soft", ""}
ALLOWED_BODY = {"full", "medium", "light"}
ALLOWED_SWEETNESS = {"dry", "off-dry", "sweet"}
YOUTUBE_PREFIXES = (
    "https://www.youtube.com/watch?v=",
    "https://youtu.be/",
)


def normalize_header(value: Any) -> str:
    return str(value or "").strip().lower()


def normalize_cell(value: Any) -> str:
    return "" if value is None else str(value).strip()


def read_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    wb = load_workbook(xlsx_path, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("Excel is empty.")

    headers = [normalize_header(v) for v in rows[0]]
    col_map = {h: idx for idx, h in enumerate(headers) if h}

    missing = [c for c in REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}")

    accepted_columns = REQUIRED_COLUMNS + OPTIONAL_COLUMNS
    result: list[dict[str, Any]] = []
    seen_sku: set[str] = set()

    for row_idx, row in enumerate(rows[1:], start=2):
        record: dict[str, Any] = {}
        is_empty_row = True

        for col in accepted_columns:
            if col not in col_map:
                continue
            value = row[col_map[col]]
            text = normalize_cell(value)
            if text:
                is_empty_row = False
            record[col] = text

        if is_empty_row:
            continue

        for col in REQUIRED_NON_EMPTY_COLUMNS:
            if not record.get(col):
                raise ValueError(f"Row {row_idx}: required field '{col}' is empty.")

        sku = str(record["sku"]).strip()
        if not sku:
            raise ValueError(f"Row {row_idx}: sku is empty.")

        # Enforce user-requested rule: id equals sku.
        record["id"] = sku

        sku_key = sku.lower()
        if sku_key in seen_sku:
            raise ValueError(f"Row {row_idx}: duplicate sku '{sku}'.")
        seen_sku.add(sku_key)

        color = str(record["color"]).strip().lower()
        acid = str(record["acid"]).strip().lower()
        tannin = str(record["tannin"]).strip().lower()
        body = str(record["body"]).strip().lower()
        sweetness = str(record["sweetness"]).strip().lower()

        if color not in ALLOWED_COLORS:
            raise ValueError(f"Row {row_idx}: invalid color '{record['color']}'.")
        if acid not in ALLOWED_ACID:
            raise ValueError(f"Row {row_idx}: invalid acid '{record['acid']}'.")
        if tannin not in ALLOWED_TANNIN:
            raise ValueError(f"Row {row_idx}: invalid tannin '{record['tannin']}'.")
        if body not in ALLOWED_BODY:
            raise ValueError(f"Row {row_idx}: invalid body '{record['body']}'.")
        if sweetness not in ALLOWED_SWEETNESS:
            raise ValueError(f"Row {row_idx}: invalid sweetness '{record['sweetness']}'.")

        try:
            record["price"] = float(str(record["price"]).replace("$", ""))
        except ValueError as exc:
            raise ValueError(f"Row {row_idx}: invalid price '{record['price']}'.") from exc

        try:
            record["vintage"] = int(float(str(record["vintage"])))
        except ValueError as exc:
            raise ValueError(f"Row {row_idx}: invalid vintage '{record['vintage']}'.") from exc

        if color in {"white", "rose", "sparkling"} and tannin:
            raise ValueError(
                f"Row {row_idx}: tannin should be empty for color '{record['color']}'."
            )

        video_url = normalize_cell(record.get("video_url", ""))
        if video_url and not video_url.startswith(YOUTUBE_PREFIXES):
            raise ValueError(
                f"Row {row_idx}: unsupported video_url '{video_url}'. "
                "Only YouTube URLs are accepted."
            )

        record["color"] = color
        record["acid"] = acid
        record["tannin"] = tannin or None
        record["body"] = body
        record["sweetness"] = sweetness
        record["video_url"] = video_url
        record["video_title"] = normalize_cell(record.get("video_title", ""))
        result.append(record)

    if not result:
        raise ValueError("No valid catalog rows found in Excel.")

    return result


WINERY_SHEET_NAME = "winery"
WINERY_REQUIRED_COLUMNS = ["id", "name_zh", "name_en"]
WINERY_OPTIONAL_COLUMNS = ["intro_zh", "intro_en", "selling_points_zh", "selling_points_en"]


def read_winery_rows(xlsx_path: Path) -> list[dict[str, Any]]:
    """Read sheet 'winery' from the same Excel. Columns: id, name_zh, name_en, intro_zh, intro_en, selling_points_zh, selling_points_en."""
    wb = load_workbook(xlsx_path, data_only=True)
    if WINERY_SHEET_NAME not in wb.sheetnames:
        return []
    ws = wb[WINERY_SHEET_NAME]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(v) for v in rows[0]]
    col_map = {h: idx for idx, h in enumerate(headers) if h}
    missing = [c for c in WINERY_REQUIRED_COLUMNS if c not in col_map]
    if missing:
        raise ValueError(f"Winery sheet missing required columns: {', '.join(missing)}")
    accepted = WINERY_REQUIRED_COLUMNS + WINERY_OPTIONAL_COLUMNS
    result: list[dict[str, Any]] = []
    for row_idx, row in enumerate(rows[1:], start=2):
        record: dict[str, Any] = {}
        for col in accepted:
            if col not in col_map:
                continue
            value = row[col_map[col]]
            record[col] = normalize_cell(value)
        if not record.get("id") and not record.get("name_zh") and not record.get("name_en"):
            continue
        record["id"] = str(record.get("id") or "").strip() or f"winery_{row_idx}"
        if not record.get("name_zh"):
            record["name_zh"] = record.get("name_en") or ""
        if not record.get("name_en"):
            record["name_en"] = record.get("name_zh") or ""
        result.append(record)
    return result


def write_json(
    products: list[dict[str, Any]],
    output_path: Path,
    wineries: list[dict[str, Any]] | None = None,
) -> None:
    payload: dict[str, Any] = {
        "version": "v1",
        "products": products,
    }
    if wineries is not None and len(wineries) > 0:
        payload["wineries"] = wineries
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--input",
        type=Path,
        default=Path("../../00_docs/docs_md/spec/fastpath/wine_catalog.xlsx"),
        help="Path to source Excel file.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path("../../00_docs/docs_md/spec/fastpath/wine_catalog.json"),
        help="Path to output JSON file.",
    )
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    base_dir = Path(__file__).resolve().parent
    input_path = (base_dir / args.input).resolve()
    output_path = (base_dir / args.output).resolve()

    products = read_rows(input_path)
    try:
        wineries = read_winery_rows(input_path)
    except ValueError as e:
        print(f"Winery sheet warning: {e}")
        wineries = []
    write_json(products, output_path, wineries=wineries if wineries else None)
    print(f"Imported {len(products)} product(s), {len(wineries)} winery(ies): {input_path} -> {output_path}")


if __name__ == "__main__":
    main()
